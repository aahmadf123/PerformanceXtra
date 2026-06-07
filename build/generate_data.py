#!/usr/bin/env python3
"""Generate the embedded dataset for PerformanceXtra's index.html.

Reads ``build/PerformanceXtra_Master_Sheet.xlsx`` and rewrites the two JSON
blobs inside ``index.html`` (the activity data and the filter taxonomy) between
their marker comments. Run this whenever the source spreadsheet changes:

    pip install openpyxl
    python3 build/generate_data.py

The app itself needs no build step or server -- index.html ships with the data
already embedded, so it works when opened directly (file://) or served.
"""
import json
import math
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
XLSX = HERE / "PerformanceXtra_Master_Sheet.xlsx"
HTML = ROOT / "index.html"

MASTER_SHEET = "Master Repository"
OPTIONS_SHEET = "Dropdown Options"


def clean(value):
    """Trim a cell value; return None for blanks."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_week(progression):
    if not progression:
        return None
    match = re.search(r"week\s+(\d+)", progression, re.I)
    return int(match.group(1)) if match else None


def parse_minutes(time_text):
    if not time_text:
        return None
    match = re.search(r"(\d+)\s*min", time_text, re.I)
    return int(match.group(1)) if match else None


def header_map(ws):
    """Map header label -> column index from row 1 (robust to column reorders)."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        label = clean(ws.cell(row=1, column=col).value)
        if label:
            headers[label] = col
    return headers


def extract_activities(ws):
    cols = header_map(ws)

    def cell(row, label):
        return ws.cell(row=row, column=cols[label])

    activities = []
    missing_links = 0
    for row in range(2, ws.max_row + 1):
        if all(ws.cell(row=row, column=c).value is None for c in range(1, ws.max_column + 1)):
            continue
        name = clean(cell(row, "Activity Name").value)
        if not name:
            continue

        # Link: prefer the cell's hyperlink target (cleaner than display text).
        link_cell = cell(row, "Link/File")
        link = None
        if link_cell.hyperlink and link_cell.hyperlink.target:
            link = link_cell.hyperlink.target.strip()
        else:
            text = clean(link_cell.value)
            if text and re.match(r"https?://", text, re.I):
                link = text
        if not link:
            missing_links += 1

        subtopics = []
        for label in ("Subtopic 1", "Subtopic 2"):
            sub = clean(cell(row, label).value)
            if sub and sub not in subtopics:
                subtopics.append(sub)

        progression = clean(cell(row, "Progression").value)
        week = parse_week(progression)
        month = math.ceil(week / 4) if week else None
        time_text = clean(cell(row, "Time").value)

        activities.append({
            "id": clean(cell(row, "Activity ID").value),
            "name": name,
            "topic": clean(cell(row, "Topic").value),
            "subtopics": subtopics,
            "type": clean(cell(row, "Content Type").value),
            "week": week,
            "progression": progression,
            "month": month,
            "frequency": clean(cell(row, "Frequency").value),
            "time": time_text,
            "timeMinutes": parse_minutes(time_text),
            "link": link,
            "instructions": clean(cell(row, "Instructions").value),
            "reflection": clean(cell(row, "Reflection Prompt").value),
        })
    return activities, missing_links


def column_values(ws, label):
    """Ordered, de-duplicated, non-blank values under every column whose row-1
    header equals ``label``."""
    values = []
    for col in range(1, ws.max_column + 1):
        if clean(ws.cell(row=1, column=col).value) == label:
            for row in range(2, ws.max_row + 1):
                value = clean(ws.cell(row=row, column=col).value)
                if value and value not in values:
                    values.append(value)
    return values


def build_months(progressions):
    weeks = sorted({parse_week(p) for p in progressions if parse_week(p)})
    if not weeks:
        return []
    max_week = max(weeks)
    months = []
    for n in range(1, math.ceil(max_week / 4) + 1):
        span = [w for w in range((n - 1) * 4 + 1, n * 4 + 1) if w <= max_week]
        if not span:
            continue
        if len(span) > 1:
            label = f"Month {n} (Weeks {span[0]}-{span[-1]})"
        else:
            label = f"Month {n} (Week {span[0]})"
        months.append({"label": label, "value": n, "weeks": span})
    return months


def extract_taxonomy(ws):
    progressions = column_values(ws, "Progression")
    return {
        "topics": column_values(ws, "Topics"),
        "subtopics": column_values(ws, "Subtopic 1 & 2 Options"),
        "types": column_values(ws, "Content Types"),
        "progressions": progressions,
        "frequencies": column_values(ws, "Frequency"),
        "months": build_months(progressions),
    }


def splice(html, start, end, payload):
    pattern = re.compile(re.escape(start) + r".*?" + re.escape(end), re.S)
    if not pattern.search(html):
        sys.exit(f"ERROR: markers {start} / {end} not found in index.html")
    replacement = f"{start}\n{payload}\n{end}"
    # Use a function replacement so backslash sequences in the JSON payload
    # (e.g. "\n" inside strings) are inserted literally, not interpreted.
    return pattern.sub(lambda _match: replacement, html, count=1)


def main():
    if not XLSX.exists():
        sys.exit(f"ERROR: source spreadsheet not found at {XLSX}")
    if not HTML.exists():
        sys.exit(f"ERROR: index.html not found at {HTML}")

    wb = openpyxl.load_workbook(XLSX)
    activities, missing_links = extract_activities(wb[MASTER_SHEET])
    taxonomy = extract_taxonomy(wb[OPTIONS_SHEET])

    # Dump JSON; neutralise any literal "</" so the data cannot terminate the
    # surrounding <script> tag early (\/ is a valid JSON escape for /).
    data_json = json.dumps(activities, ensure_ascii=False, indent=2).replace("</", "<\\/")
    tax_json = json.dumps(taxonomy, ensure_ascii=False, indent=2).replace("</", "<\\/")

    html = HTML.read_text(encoding="utf-8")
    html = splice(html, "<!-- PX_DATA_START -->", "<!-- PX_DATA_END -->", data_json)
    html = splice(html, "<!-- PX_TAXONOMY_START -->", "<!-- PX_TAXONOMY_END -->", tax_json)
    HTML.write_text(html, encoding="utf-8")

    print(f"Wrote {HTML.relative_to(ROOT)}")
    print(f"  Activities:    {len(activities)}")
    print(f"  Missing links: {missing_links}")
    print(f"  Topics:        {len(taxonomy['topics'])}")
    print(f"  Subtopics:     {len(taxonomy['subtopics'])}")
    print(f"  Content types: {len(taxonomy['types'])}")
    print(f"  Progressions:  {len(taxonomy['progressions'])}")
    print(f"  Months:        {len(taxonomy['months'])}")


if __name__ == "__main__":
    main()
